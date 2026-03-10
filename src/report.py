"""
Build Excel report from decrypted rows and password-protect the workbook.
Sheet "Fitness Data", columns A-P per ROADMAP; sort by date ascending.
Uses report_password only; encrypts file with msoffcrypto so Excel prompts to open.
"""
import io
import logging
import tempfile
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

HEADERS = [
    "date", "weight", "fat_percent", "bmi", "fat_weight", "lean_weight",
    "neck", "shoulders", "biceps", "forearms", "chest", "above_navel",
    "waist", "hips", "thighs", "calves",
]

# Map report header -> decrypted JSON key (camelCase from iOS)
_FIELD_MAP = {
    "date": "date",
    "weight": "weight",
    "fat_percent": "fatPercentage",
    "bmi": "bmi",
    "fat_weight": "fatWeight",
    "lean_weight": "leanWeight",
    "neck": "neck",
    "shoulders": "shoulders",
    "biceps": "biceps",
    "forearms": "forearms",
    "chest": "chest",
    "above_navel": "aboveNavel",  # fallback: navel
    "waist": "waist",
    "hips": "hips",
    "thighs": "thighs",
    "calves": "calves",
}


def _format_date(val) -> str | None:
    """Format ISO date for Excel display (YYYY-MM-DD)."""
    if val is None:
        return None
    s = str(val).strip()
    if "T" in s:
        return s.split("T")[0]
    return s[:10] if len(s) >= 10 else s


def _row_from_measurement(m: dict) -> list:
    """Map decrypted measurement dict to Excel row (same order as HEADERS)."""
    row = []
    for h in HEADERS:
        key = _FIELD_MAP.get(h, h)
        val = m.get(key)
        if key == "aboveNavel" and val is None:
            val = m.get("navel")
        if h == "date":
            val = _format_date(val)
        row.append(val)
    return row


def build_workbook(rows: list[dict]) -> openpyxl.Workbook:
    """Build workbook 'Fitness Data' with header row and data sorted by date ascending."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fitness Data"
    for col, name in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=name)
    # Sort by date ascending
    sorted_rows = sorted(rows, key=lambda m: _format_date(m.get("date")) or "")
    for row_idx, m in enumerate(sorted_rows, start=2):
        for col_idx, val in enumerate(_row_from_measurement(m), start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    return wb


def build_report_bytes(
    decrypted_rows: list[dict],
    report_password: str,
    user_id: int,
) -> bytes:
    """
    Build Excel from decrypted rows, encrypt with report_password (password to open),
    return file bytes. Uses msoffcrypto OOXML encryption.
    """
    wb = build_workbook(decrypted_rows)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        wb.save(tmp_path)
        from msoffcrypto.format.ooxml import OOXMLFile
        with open(tmp_path, "rb") as plain:
            file = OOXMLFile(plain)
            encrypted = io.BytesIO()
            file.encrypt(report_password, encrypted)
            return encrypted.getvalue()
    finally:
        Path(tmp_path).unlink(missing_ok=True)
